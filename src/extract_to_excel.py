"""
EXTRACT TO EXCEL: PDF se important tender fields nikal ke Excel sheet mein bharna
Diagram ka extension: Retrieval + LLM ab structured JSON fields extract karte hain
(manual Q&A ki jagah)

Run: python extract_to_excel.py
"""
"""
EXTRACT TO EXCEL: PDF se important tender fields nikal ke Excel sheet mein bharna
"""
import os
import re
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import PDF_PATH, TOP_K, FAISS_INDEX_PATH, CHUNKS_CACHE_PATH


def ensure_index_built():
    if os.path.exists(FAISS_INDEX_PATH) and os.path.exists(CHUNKS_CACHE_PATH):
        print("Existing index mil gaya — dobara PDF process nahi karunga.\n")
        return
    print("Index nahi mila — pehli baar PDF process kar raha hoon...\n")
    from step3_vector_db import build_and_save
    build_and_save()
    print("\nIndex ban gaya aur save ho gaya.\n")


from step4_semantic_search import load_index_and_chunks, build_bm25_index, hybrid_search
from step5_llm_generate import call_llm

EXCEL_OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "tender_extracted_data.xlsx")

FIELD_GROUPS = {
    "identifiers_and_dates": {
        "query": "bid number tender number GEM bid ID dated bid end date bid opening date time",
        "fields": {
            "tender_reference_number": "The unique Bid/Tender/EOI reference number or ID (may be labeled 'Bid Number', 'Tender Number', or 'EOI Number', e.g. GEM/2025/B/xxxxxx)",
            "bid_end_date": "Bid End Date and Time (as written in document)",
            "bid_opening_date": "Bid Opening Date and Time (as written in document)",
        }
    },
    "organisation_details": {
        "query": "name of organisation issuing tender ministry department office address for receiving bids buyer email item category total quantity",
        "fields": {
            "organisation_name": "Name of the organisation/company issuing this tender",
            "ministry_name": "Ministry/State Name",
            "department_name": "Department Name",
            "office_name": "Office Name",
            "receivers_address": "Address where bids/tender documents should be sent (receiver's address)",
            "buyer_email": "Buyer's email address",
            "total_quantity": "Total Quantity being tendered (numeric only)",
            "item_category": "Item Category / description of item(s) being tendered",
        }
    },
    "financial": {
        "query": "EMD earnest money deposit amount PBG performance bank guarantee percentage security deposit tender fee processing fee tender value",
        "fields": {
            "emd_required": "Is EMD (Earnest Money Deposit) required? (Yes/No)",
            "emd_amount": "EMD amount (numeric value only, no currency symbol)",
            "emd_mode": "Mode of EMD payment (e.g. Demand Draft, Bank Guarantee)",
            "pbg_required": "Is PBG (Performance Bank Guarantee) required? (Yes/No)",
            "pbg_percentage": "PBG percentage (numeric only)",
            "pbg_duration": "PBG duration in days/months (numeric only)",
            "sd_required": "Is Security Deposit required? (Yes/No)",
            "sd_percentage": "Security Deposit percentage (numeric only)",
            "sd_duration": "Security Deposit duration (numeric only)",
            "tender_fee_required": "Is Tender Fee required? (Yes/No)",
            "tender_fee_amount": "Tender fee amount (numeric only)",
            "processing_fee_required": "Is processing fee required? (Yes/No)",
            "processing_fee_amount": "Processing fee amount (numeric only)",
            "tender_value": "Total tender/project value (numeric only)",
        }
    },
    "eligibility": {
        "query": "eligibility criteria average annual turnover working capital solvency certificate net worth technical experience OEM",
        "fields": {
            "avg_annual_turnover_value": "Required average annual turnover value (numeric only)",
            "working_capital_value": "Required working capital value (numeric only)",
            "solvency_certificate_value": "Required solvency certificate value (numeric only)",
            "net_worth_value": "Required net worth value (numeric only)",
            "technical_eligibility_age": "Years of technical experience required (numeric only)",
            "oem_experience": "OEM experience requirement (short text)",
            "custom_eligibility_criteria": "Any other unique eligibility criteria (short summary)",
        }
    },
    "timelines_and_terms": {
        "query": "bid validity period delivery time supply installation payment terms liquidated damages last date submission",
        "fields": {
            "bid_validity_days": "Bid validity period in days (numeric only)",
            "delivery_time_supply": "Delivery time for supply in days (numeric only)",
            "delivery_time_installation_days": "Delivery time for installation in days (numeric only)",
            "payment_terms_supply": "Payment terms for supply, percentage (numeric only)",
            "payment_terms_installation": "Payment terms for installation, percentage (numeric only)",
            "max_ld_percentage": "Maximum Liquidated Damages percentage (numeric only)",
            "ld_percentage_per_week": "LD percentage charged per week (numeric only)",
            "physical_docs_required": "Are physical documents required to be submitted? (Yes/No)",
            "physical_docs_deadline": "Deadline for physical document submission (date, if mentioned)",
        }
    },
    "process_and_evaluation": {
        "query": "commercial evaluation criteria manufacturer authorization form reverse auction bid opening",
        "fields": {
            "commercial_evaluation": "How is commercial evaluation done? (short text, e.g. L1 basis)",
            "maf_required": "Is Manufacturer Authorization Form (MAF) required? (Yes/No)",
            "reverse_auction_applicable": "Is reverse auction applicable? (Yes/No)",
        }
    },
    "courier_contact": {
        "query": "address for submission of bid courier contact phone email communication",
        "fields": {
            "courier_name": "Contact person name for bid submission",
            "courier_phone": "Contact phone number",
            "courier_address_line_1": "Address line 1 for bid/courier submission",
            "courier_city": "City for bid submission address",
            "courier_state": "State for bid submission address",
            "courier_pincode": "Pincode for bid submission address",
        }
    },
}


def clean_json_response(text):
    text = re.sub(r"```json\s*|```\s*", "", text).strip()
    match = re.search(r"\{.*\}", text, re.DOTALL)
    return match.group(0) if match else text


def build_extraction_prompt(fields_dict, context):
    field_descriptions = "\n".join([f'- "{key}": {desc}' for key, desc in fields_dict.items()])
    prompt = f"""You are a data extraction assistant. Extract the following fields from the context below.
Respond with ONLY a valid JSON object (no markdown, no explanation) using EXACTLY these keys:

{field_descriptions}

Rules:
- If a value is not found in the context, use null (not the string "null", actual JSON null)
- For numeric fields, return only the number (no currency symbols, no commas, no units)
- For Yes/No fields, return exactly "Yes" or "No" or null if unclear
- Keep text fields concise (under 15 words)

Context:
{context}

Respond with ONLY the JSON object:"""
    return prompt


def extract_field_group(group_name, group_config, index, chunks, bm25):
    print(f"Extracting group: {group_name}...")
    retrieved = hybrid_search(group_config["query"], index, chunks, bm25, top_k=TOP_K)
    context = "\n\n".join([f"[Page {c['page']}]: {c['content']}" for c in retrieved])
    prompt = build_extraction_prompt(group_config["fields"], context)
    response = call_llm(prompt)
    try:
        cleaned = clean_json_response(response)
        data = json.loads(cleaned)
    except (json.JSONDecodeError, AttributeError):
        print(f"   JSON parse failed for {group_name}")
        data = {key: None for key in group_config["fields"]}
    return data


def extract_all_fields():
    index, chunks = load_index_and_chunks()
    bm25 = build_bm25_index(chunks)
    all_data = {}
    for group_name, group_config in FIELD_GROUPS.items():
        group_data = extract_field_group(group_name, group_config, index, chunks, bm25)
        all_data.update(group_data)
    all_data["source_pdf"] = os.path.basename(PDF_PATH)
    return all_data


def save_to_excel(data, output_path=EXCEL_OUTPUT_PATH):
    headers = list(data.keys())
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
        existing_headers = [cell.value for cell in ws[1]]
        if existing_headers != headers:
            wb = Workbook()
            ws = wb.active
            _write_headers(ws, headers)
    else:
        wb = Workbook()
        ws = wb.active
        _write_headers(ws, headers)

    row = [data.get(h, "") if data.get(h) is not None else "" for h in headers]
    ws.append(row)
    wb.save(output_path)
    print(f"\nSaved to: {output_path}")


def _write_headers(ws, headers):
    ws.title = "Tender Data"
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill


if __name__ == "__main__":
    ensure_index_built()
    print(f"Processing: {PDF_PATH}\n")
    data = extract_all_fields()
    print("\nEXTRACTED DATA:")
    for key, value in data.items():
        print(f"  {key}: {value}")
    save_to_excel(data)
    print("\nExtraction complete!")